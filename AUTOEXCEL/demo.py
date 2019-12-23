from openpyxl import load_workbook

wb = load_workbook(r"C:\Users\liuyang\Desktop\import_template\运单\运单.xlsx") # 打开运单
# wb_d = load_workbook('./订单.xlsx')
# wb_q = load_workbook('./清单.xlsx')
sheets = wb.sheetnames # 得到所有的sheet
print(sheets)
ws = wb[sheets[0]] # 取第一个sheet的名字设为ws
print(ws)

# 打印第二行的所有数据

# 订单的（电商企业代码、电商企业名称）= 清单（申报企业代码、申报企业名称）
# 定义电商企业代码为ebp_code、名称为ebp_name



# 运单的（物流企业代码、物流企业名称）= 清单（物流企业代码、物流企业名称）

